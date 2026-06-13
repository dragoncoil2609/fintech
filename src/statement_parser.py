"""Bank-statement parsing — CSV and Excel (.xlsx/.xls).

Vietnamese bank Excel exports vary wildly (column names, header row position,
single signed "Số tiền" vs separate "Ghi nợ"/"Ghi có"). Rather than a parser per
bank, we detect columns by a synonym dictionary that's trivial to extend. A
debit/credit pair is collapsed into one signed amount (credit = income +, debit =
expense −). Unknown layouts raise a clear error (the column-mapping UI / `mapping`
override is the escape hatch).

Returns rows shaped like the CSV parser: [{"date","description","amount"}].
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime

from . import handlers

_SYNONYMS: dict[str, list[str]] = {
    "date": [
        "ngay", "ngay giao dich", "ngay gd", "transaction date", "date", "trans date",
        "thoi gian", "ngay hieu luc", "posting date", "value date", "ngay ghi so",
    ],
    "description": [
        "dien giai", "noi dung", "noi dung giao dich", "mo ta", "description",
        "details", "ghi chu", "remark", "transaction detail", "memo", "noi dung ck",
    ],
    "amount": [
        "so tien", "amount", "gia tri", "so tien gd", "transaction amount", "value",
    ],
    "debit": [
        "ghi no", "no", "debit", "tien ra", "phat sinh no", "withdrawal", "money out",
        "rut", "debit amount",
    ],
    "credit": [
        "ghi co", "co", "credit", "tien vao", "phat sinh co", "deposit", "money in",
        "nop", "credit amount",
    ],
}


def _strip_accents(text: str) -> str:
    nfkd = unicodedata.normalize("NFD", text)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def _norm(cell) -> str:
    return _strip_accents(str(cell or "")).lower().strip()


def _match_column(header_cell: str) -> str | None:
    h = _norm(header_cell)
    if not h:
        return None
    for field, syns in _SYNONYMS.items():
        for s in syns:
            if h == s or h.startswith(s) or s in h:
                return field
    return None


def _parse_amount(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    neg = s.startswith("-") or s.startswith("(")
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    amount = float(digits)
    return -amount if neg else amount


def _norm_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d",
                "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return None


def _detect_header(rows: list[list]) -> tuple[int, dict[str, int]] | None:
    """Find the header row (within the first 20) and map field → column index."""
    for i, row in enumerate(rows[:20]):
        mapping: dict[str, int] = {}
        for j, cell in enumerate(row):
            field = _match_column(cell)
            if field and field not in mapping:
                mapping[field] = j
        has_date = "date" in mapping
        has_money = "amount" in mapping or "debit" in mapping or "credit" in mapping
        if has_date and has_money:
            return i, mapping
    return None


def parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        detected = _detect_header(rows)
        if not detected:
            continue
        header_idx, cols = detected
        return _rows_from_table(rows[header_idx + 1:], cols)
    raise ValueError(
        "Không nhận diện được cột trong file Excel. Cần có cột Ngày và "
        "Số tiền (hoặc Ghi nợ/Ghi có)."
    )


def parse_xls(content: bytes) -> list[dict]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError(
            ".xls cũ chưa được hỗ trợ (cần: pip install xlrd). "
            "Hãy lưu lại dạng .xlsx hoặc .csv."
        ) from exc
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    detected = _detect_header(rows)
    if not detected:
        raise ValueError("Không nhận diện được cột trong file .xls.")
    header_idx, cols = detected
    return _rows_from_table(rows[header_idx + 1:], cols)


def _rows_from_table(data_rows: list[list], cols: dict[str, int]) -> list[dict]:
    out: list[dict] = []
    for row in data_rows:
        if not row or not any(c not in (None, "") for c in row):
            continue
        d = _norm_date(_cell(row, cols.get("date")))
        if not d:
            continue
        if "amount" in cols:
            amount = _parse_amount(_cell(row, cols["amount"]))
        else:
            debit = _parse_amount(_cell(row, cols.get("debit"))) or 0
            credit = _parse_amount(_cell(row, cols.get("credit"))) or 0
            amount = credit - abs(debit)
        if amount is None or amount == 0:
            continue
        desc = str(_cell(row, cols.get("description")) or "").strip()
        out.append({"date": d, "description": desc, "amount": float(amount)})
    return out


def _cell(row: list, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_statement(content: bytes, filename: str) -> list[dict]:
    """Dispatch by extension/signature → normalised transaction rows."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or content[:4] == b"PK\x03\x04":
        return parse_xlsx(content)
    if name.endswith(".xls") or content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return parse_xls(content)
    return handlers._parse_csv(content)
